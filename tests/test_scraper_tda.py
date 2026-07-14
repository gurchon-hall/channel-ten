"""Tests for the smeea/vdb TDA read-only source (channel_ten.scraper._tda)."""

import logging
import zipfile
from datetime import date, datetime
from io import BytesIO
from typing import Any
from unittest.mock import MagicMock

import httpx
import pytest
from openpyxl import Workbook

from channel_ten.scraper._tda import (
    fetch_tda_archive,
    iter_tda_deck_texts,
    list_tda_archive_ids,
    parse_archon_xlsx,
    read_archon_xlsx,
)


def _mock_response(
    status_code: int = 200,
    json_data: Any = None,
    content: bytes = b"",
    headers: dict[str, str] | None = None,
) -> MagicMock:
    resp = MagicMock()
    resp.status_code = status_code
    resp.headers = headers or {}
    resp.content = content
    resp.json.return_value = json_data
    resp.raise_for_status = MagicMock()
    return resp


def _build_archon_xlsx(
    name: str = "Finnish Nationals 2022",
    location: str = "Espoo - Finland",
    event_date: datetime = datetime(2022, 11, 5),
    rounds: int = 4,
    players: int = 45,
    winner_name: str = "Teemu Sainomaa",
    winner_seat: int = 41,
    winner_vekn: int = 3070069,
    methuselah_seat: int | None = None,
) -> bytes:
    """Build a minimal, structurally-faithful archon.xlsx in memory."""
    wb = Workbook()
    info = wb.active
    info.title = "Tournament Info"
    info.append([name])
    info.append(["Event Name:", name])
    info.append(["Event Date (DD-MON-YY):", event_date])
    info.append(["City:", location])
    info.append(["Number of Players:", players])
    info.append(["Number of Rounds (including final):", rounds])

    standings = wb.create_sheet("Standings")
    standings.append([name])
    standings.append(["Final Standings"])
    standings.append([None, True])
    standings.append(["Final Rank", "Target Rank", "#", "Name", "Prelim GWs"])
    standings.append([1, 1, winner_seat, winner_name, 2])
    standings.append([2, 2, 1, "Runner Up", 2])

    meth = wb.create_sheet("Methuselahs")
    meth.append([name])
    meth.append([None] * 15 + ["champ"])
    meth.append(["Player Information"])
    meth.append([players, "players entered"])
    meth.append(["Player", "Player", "Player", "Player", "V:EKN"])
    meth.append(["Num.", "First Name", "Last Name", "City", "Num."])
    meth.append([methuselah_seat or winner_seat, "Teemu", "Sainomaa", None, winner_vekn])
    meth.append([1, "Someone", "Else", None, 1003636])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_archive_zip(deck_texts: dict[str, str], xlsx_bytes: bytes | None = None) -> bytes:
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("archon.xlsx", xlsx_bytes if xlsx_bytes is not None else _build_archon_xlsx())
        for filename, text in deck_texts.items():
            zf.writestr(filename, text)
    return buf.getvalue()


_SAMPLE_DECK_TEXT = """Deck Name: Test Deck
Author: 3070069
Description: A test deck.

Crypt (2 cards, min=4 max=4 avg=4.00)
======================================
2x Nathan Turner  4 PRO ani  Gangrel:6

Library (1 cards)
==================
Master (1)
-----------
1x Blood Doll
"""


# ---------------------------------------------------------------------------
# list_tda_archive_ids
# ---------------------------------------------------------------------------


class TestListTdaArchiveIds:
    def test_filters_and_sorts_archive_ids(self):
        tree = {
            "tree": [
                {"type": "blob", "path": "frontend/public/tournaments/10367.zip"},
                {"type": "blob", "path": "frontend/public/tournaments/online2.zip"},
                {"type": "blob", "path": "frontend/public/tournaments/online1.zip"},
                {"type": "blob", "path": "README.md"},
                {"type": "tree", "path": "frontend/public/tournaments"},
                {"type": "blob", "path": "frontend/public/tournaments/9999.zip"},
            ],
            "truncated": False,
        }
        client = MagicMock()
        client.get.return_value = _mock_response(json_data=tree)

        result = list_tda_archive_ids(client, token=None, delay=0)
        assert result == ["10367", "9999", "online1", "online2"]

    def test_rate_limit_raises_runtime_error(self):
        client = MagicMock()
        client.get.return_value = _mock_response(
            status_code=403, headers={"X-RateLimit-Remaining": "0"}
        )
        with pytest.raises(RuntimeError, match="rate limit"):
            list_tda_archive_ids(client, token=None, delay=0)

    def test_truncated_tree_logs_warning(self, caplog: pytest.LogCaptureFixture):
        tree = {
            "tree": [{"type": "blob", "path": "frontend/public/tournaments/1.zip"}],
            "truncated": True,
        }
        client = MagicMock()
        client.get.return_value = _mock_response(json_data=tree)

        with caplog.at_level(logging.WARNING):
            result = list_tda_archive_ids(client, token=None, delay=0)
        assert result == ["1"]
        assert any("truncated" in r.message for r in caplog.records)


# ---------------------------------------------------------------------------
# fetch_tda_archive
# ---------------------------------------------------------------------------


class TestFetchTdaArchive:
    def test_returns_bytes_on_200(self):
        client = MagicMock()
        client.get.return_value = _mock_response(status_code=200, content=b"PK\x03\x04zip")
        assert fetch_tda_archive(client, "10367", delay=0) == b"PK\x03\x04zip"

    def test_returns_none_on_404(self):
        client = MagicMock()
        client.get.return_value = _mock_response(status_code=404)
        assert fetch_tda_archive(client, "10367", delay=0) is None

    def test_raises_on_500(self):
        client = MagicMock()
        resp = _mock_response(status_code=500)
        resp.raise_for_status.side_effect = httpx.HTTPStatusError(
            "boom", request=MagicMock(), response=MagicMock()
        )
        client.get.return_value = resp
        with pytest.raises(httpx.HTTPStatusError):
            fetch_tda_archive(client, "10367", delay=0)

    def test_requests_expected_raw_url(self):
        client = MagicMock()
        client.get.return_value = _mock_response(status_code=200, content=b"x")
        fetch_tda_archive(client, "online1", delay=0)
        args, _ = client.get.call_args
        assert args[0].endswith("/tournaments/online1.zip")


# ---------------------------------------------------------------------------
# parse_archon_xlsx
# ---------------------------------------------------------------------------


class TestParseArchonXlsx:
    def test_extracts_event_metadata(self):
        meta = parse_archon_xlsx(_build_archon_xlsx())
        assert meta.name == "Finnish Nationals 2022"
        assert meta.location == "Espoo - Finland"
        assert meta.date_start == date(2022, 11, 5)
        assert meta.rounds_format == "3R+F"
        assert meta.players_count == 45

    def test_resolves_winner_and_vekn_number_via_seat(self):
        meta = parse_archon_xlsx(_build_archon_xlsx())
        assert meta.winner == "Teemu Sainomaa"
        assert meta.winner_vekn_number == 3070069

    def test_winner_vekn_number_none_when_seat_unresolvable(self):
        # Standings seat (41) has no matching row in Methuselahs (seat 7 instead).
        xlsx = _build_archon_xlsx(winner_seat=41, methuselah_seat=7)
        meta = parse_archon_xlsx(xlsx)
        assert meta.winner == "Teemu Sainomaa"
        assert meta.winner_vekn_number is None


# ---------------------------------------------------------------------------
# read_archon_xlsx / iter_tda_deck_texts
# ---------------------------------------------------------------------------


class TestArchiveContents:
    def test_read_archon_xlsx_returns_bytes(self):
        xlsx_bytes = _build_archon_xlsx()
        archive = _build_archive_zip({"a.txt": _SAMPLE_DECK_TEXT}, xlsx_bytes)
        assert read_archon_xlsx(archive) == xlsx_bytes

    def test_iter_tda_deck_texts_skips_xlsx(self):
        archive = _build_archive_zip({"a.txt": _SAMPLE_DECK_TEXT, "b.txt": _SAMPLE_DECK_TEXT})
        results = list(iter_tda_deck_texts(archive))
        names = {name for name, _ in results}
        assert names == {"a.txt", "b.txt"}
        assert all("Author: 3070069" in text for _, text in results)

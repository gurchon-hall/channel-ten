"""Read-only source for VDB Tournament Deck Archives (TDA).

No official VEKN archive of full tournament decklists (every participant, not
just the winner) exists. The closest available source is
[`smeea/vdb`](https://github.com/smeea/vdb)'s
``frontend/public/tournaments/*.zip`` archives: one zip per tournament,
containing the standard VEKN "Archon" tournament-report spreadsheet
(``archon.xlsx``) plus one ``.txt`` per participant's deck, in the same
textual format as a TWD post's deck block.

Zip filenames are the archive id: numeric when the event has a VEKN calendar
id (e.g. ``10367.zip``), or a short slug for recurring online events that
never got one (e.g. ``online1.zip``). This module only reads the archive; it
never fetches from ``vekn.net`` — event metadata (name, date, location,
rounds, players, winner) comes entirely from ``archon.xlsx``, which is the
same report organizers already submit to VEKN.

GitHub access differs from :mod:`channel_ten.scraper._twda` in one respect:
  - Listing uses the contents API scoped to the tournaments folder, not the
    recursive git trees API :mod:`channel_ten.scraper._twda` uses — unlike
    GiottoVerducci/TWD (a repo dedicated to deck files), smeea/vdb is a full web
    app whose recursive tree is ~15 MB of mostly unrelated frontend assets, and
    GitHub intermittently 500s trying to serve it. The folder-scoped contents API
    returns only the tournaments folder's own entries (a few KB) and has been
    reliable in comparison. Pass a token to raise the 60/hour unauthenticated
    rate limit either way.
  - Per-archive fetches use raw.githubusercontent.com, not subject to that limit.
"""

import logging
import re
import time
import zipfile
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO

import httpx
from openpyxl import load_workbook

from channel_ten.scraper._http import DEFAULT_DELAY_SECONDS

logger = logging.getLogger(__name__)

VDB_OWNER = "smeea"
VDB_REPO = "vdb"
VDB_BRANCH = "master"
VDB_TOURNAMENTS_FOLDER = "frontend/public/tournaments"

_GITHUB_API = "https://api.github.com"
VDB_RAW_BASE = f"https://raw.githubusercontent.com/{VDB_OWNER}/{VDB_REPO}/{VDB_BRANCH}/{VDB_TOURNAMENTS_FOLDER}"

# Matches archive blob paths like "frontend/public/tournaments/10367.zip" and
# "frontend/public/tournaments/online1.zip", capturing the archive id.
_ARCHIVE_PATH_RE = re.compile(rf"^{VDB_TOURNAMENTS_FOLDER}/([^/]+)\.zip$")

_TOURNAMENT_INFO_SHEET = "Tournament Info"
_STANDINGS_SHEET = "Standings"
_METHUSELAHS_SHEET = "Methuselahs"


def _tda_headers(token: str | None = None) -> dict[str, str]:
    """Build GitHub API headers. Authentication is optional but raises the limit."""
    headers = {
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def list_tda_archive_ids(
    client: httpx.Client,
    token: str | None = None,
    delay: float = DEFAULT_DELAY_SECONDS,
) -> list[str]:
    """Return the sorted archive ids of every ``tournaments/<id>.zip`` file in VDB.

    Uses the contents API scoped to ``VDB_TOURNAMENTS_FOLDER`` (one request) rather
    than the recursive git trees API — see the module docstring for why. This is
    limited to 1,000 entries per GitHub's contents API; smeea/vdb has a few dozen
    archives as of this writing, far below that.

    Raises:
        RuntimeError: when the request is rejected because the API rate limit has
            been exhausted (advises setting a token).
        httpx.HTTPStatusError: for other unexpected HTTP errors.
    """
    url = f"{_GITHUB_API}/repos/{VDB_OWNER}/{VDB_REPO}/contents/{VDB_TOURNAMENTS_FOLDER}"
    logger.debug("GET %s", url)
    resp = client.get(url, headers=_tda_headers(token), params={"ref": VDB_BRANCH})

    if resp.status_code == 401:
        raise RuntimeError(
            "GitHub API returned 401 Unauthorized while listing smeea/vdb. "
            "Your GITHUB_TOKEN may be expired or invalid — unset it or pass a valid "
            "token via --github-token."
        )
    if resp.status_code == 403 and resp.headers.get("X-RateLimit-Remaining") == "0":
        raise RuntimeError(
            "GitHub API rate limit exceeded while listing smeea/vdb. "
            "Set the GITHUB_TOKEN environment variable (or pass --github-token) "
            "to raise the limit."
        )
    resp.raise_for_status()
    time.sleep(delay)

    data = resp.json()
    archive_ids: list[str] = []
    for entry in data:
        if entry.get("type") != "file":
            continue
        match = _ARCHIVE_PATH_RE.match(entry.get("path", ""))
        if match:
            archive_ids.append(match.group(1))

    return sorted(archive_ids)


def fetch_tda_archive(
    client: httpx.Client,
    archive_id: str,
    delay: float = DEFAULT_DELAY_SECONDS,
) -> bytes | None:
    """Fetch the raw zip bytes for *archive_id*, or ``None`` if it does not exist.

    Returns the zip content on HTTP 200, ``None`` on 404, and raises on other
    HTTP errors.
    """
    url = f"{VDB_RAW_BASE}/{archive_id}.zip"
    logger.debug("GET %s", url)
    resp = client.get(url, follow_redirects=True)
    if resp.status_code == 404:
        logger.debug("No TDA archive for id %s", archive_id)
        time.sleep(delay)
        return None
    resp.raise_for_status()
    time.sleep(delay)
    return resp.content


@dataclass
class TdaStandingRow:
    """One participant's row from archon.xlsx's ``Standings`` sheet.

    ``vekn_number`` is joined in from ``Methuselahs`` via ``seat`` (Standings' own
    ``#`` column) — Standings does not carry a VEKN number itself. ``gw`` is
    "Prelim GWs" (no separate final-round GW is tracked). ``vp`` is "Prelim VPs" +
    "Final VPs" (0 when a player did not reach the final) — matches the total each
    deck's own ``Description:`` line encodes, e.g. ``2GW8.5+3``.
    """

    final_rank: int
    target_rank: int
    seat: int
    name: str
    gw: int
    vp: float
    tp: int
    vekn_number: int | None


@dataclass
class TdaEventMeta:
    """Event-level metadata extracted from one archive's ``archon.xlsx``."""

    name: str
    location: str
    date_start: date
    rounds_format: str
    players_count: int
    winner: str
    winner_vekn_number: int | None
    standings: list[TdaStandingRow]


def _row_label_map(rows: Sequence[tuple[object, ...]]) -> dict[str, tuple[object, ...]]:
    """Map each row's first-cell text (stripped, no trailing colon) to the full row."""
    labels: dict[str, tuple[object, ...]] = {}
    for row in rows:
        if row and isinstance(row[0], str):
            labels[row[0].strip()] = row
    return labels


def _normalize_rounds(total_rounds: int) -> str:
    """archon.xlsx reports the round count *including* the final — e.g. 4 -> "3R+F"."""
    return f"{total_rounds - 1}R+F"


def _cell_int(value: object) -> int:
    """Convert an archon.xlsx numeric cell (read back as int or float) to int."""
    if isinstance(value, (int, float)):
        return int(value)
    raise ValueError(f"Expected a numeric cell, got {value!r}")


def parse_archon_xlsx(data: bytes) -> TdaEventMeta:
    """Extract event metadata from one archive's ``archon.xlsx`` report.

    Reads three sheets:
      - ``Tournament Info``: name, date, city, rounds, players count.
      - ``Standings``: every participant's final rank, GW/VP/TP; the Final Rank 1
        row gives the winner.
      - ``Methuselahs``: maps each seat number to its VEKN member number, joined
        into each Standings row (and used to resolve ``winner_vekn_number``).
    """
    wb = load_workbook(BytesIO(data), data_only=True)

    info_rows = list(wb[_TOURNAMENT_INFO_SHEET].iter_rows(values_only=True))
    info = _row_label_map(info_rows)

    name = str(info["Event Name:"][1]).strip()
    location = str(info["City:"][1]).strip()
    event_date = info["Event Date (DD-MON-YY):"][1]
    if isinstance(event_date, datetime):
        date_start = event_date.date()
    elif isinstance(event_date, date):
        date_start = event_date
    else:
        raise ValueError(f"Unparseable event date in archon.xlsx: {event_date!r}")
    rounds_format = _normalize_rounds(_cell_int(info["Number of Rounds (including final):"][1]))
    players_count = _cell_int(info["Number of Players:"][1])

    seat_to_vekn = _seat_to_vekn_map(list(wb[_METHUSELAHS_SHEET].iter_rows(values_only=True)))
    standings = _parse_standings(
        list(wb[_STANDINGS_SHEET].iter_rows(values_only=True)), seat_to_vekn
    )
    winner_row = next((row for row in standings if row.final_rank == 1), None)
    if winner_row is None:
        raise ValueError("Standings sheet has no Final Rank 1 row")

    return TdaEventMeta(
        name=name,
        location=location,
        date_start=date_start,
        rounds_format=rounds_format,
        players_count=players_count,
        winner=winner_row.name,
        winner_vekn_number=winner_row.vekn_number,
        standings=standings,
    )


def _seat_to_vekn_map(rows: Sequence[tuple[object, ...]]) -> dict[int, int]:
    """Build {seat number (Methuselahs "Num." column) -> VEKN number} from Methuselahs.

    Data rows start right after the header row whose first cell is "Num." — the
    two rows above it are grouped/merged category headers, not data.
    """
    header_idx = next(
        (i for i, row in enumerate(rows) if row and str(row[0]).strip() == "Num."),
        None,
    )
    if header_idx is None:
        return {}

    seat_map: dict[int, int] = {}
    for row in rows[header_idx + 1 :]:
        if not row or not isinstance(row[0], int) or len(row) <= 4:
            continue
        vekn_number = row[4]
        if isinstance(vekn_number, int):
            seat_map[row[0]] = vekn_number
    return seat_map


def _cell_float(value: object) -> float:
    """Convert an archon.xlsx numeric cell (read back as int or float) to float."""
    if isinstance(value, (int, float)):
        return float(value)
    raise ValueError(f"Expected a numeric cell, got {value!r}")


def _parse_standings(
    rows: Sequence[tuple[object, ...]], seat_to_vekn: dict[int, int]
) -> list[TdaStandingRow]:
    """Parse every participant's row from the Standings sheet, joined with *seat_to_vekn*.

    Column layout: Final Rank, Target Rank, # (seat), Name, Prelim GWs, Prelim VPs,
    Final VPs, TPs. ``vp`` sums Prelim VPs and Final VPs — 0 when a player did not
    reach the final (that cell is blank) — matching the total each deck's own
    ``Description:`` line encodes, e.g. ``2GW8.5+3``.
    """
    header_idx = next(
        (i for i, row in enumerate(rows) if row and row[0] == "Final Rank"),
        None,
    )
    if header_idx is None:
        raise ValueError("Standings sheet has no 'Final Rank' header row")

    standings: list[TdaStandingRow] = []
    for row in rows[header_idx + 1 :]:
        if not row or not isinstance(row[0], int) or len(row) <= 7:
            continue
        seat = _cell_int(row[2])
        final_vp = _cell_float(row[6]) if row[6] is not None else 0.0
        standings.append(
            TdaStandingRow(
                final_rank=row[0],
                target_rank=_cell_int(row[1]),
                seat=seat,
                name=str(row[3]).strip(),
                gw=_cell_int(row[4]),
                vp=_cell_float(row[5]) + final_vp,
                tp=_cell_int(row[7]),
                vekn_number=seat_to_vekn.get(seat),
            )
        )
    return standings


_DECK_FILENAME_RANK_RE = re.compile(r"_(\d+)\.txt$", re.IGNORECASE)


def target_rank_from_deck_filename(filename: str) -> int | None:
    """Extract the Standings ``Target Rank`` a deck ``.txt`` filename encodes.

    Deck filenames follow ``<slug>_<N>.txt``. ``N`` was confirmed (against a live
    archive) to be the Standings sheet's ``Target Rank`` column — not the seat
    number and not ``Final Rank`` (which ties every non-winning finalist at the same
    value, so it cannot identify one specific deck). Returns ``None`` if the filename
    doesn't match the expected pattern (e.g. an online-event archive using different
    naming).
    """
    match = _DECK_FILENAME_RANK_RE.search(filename)
    return int(match.group(1)) if match else None


def standing_for_target_rank(
    standings: Sequence[TdaStandingRow], target_rank: int
) -> TdaStandingRow | None:
    """Return the Standings row matching *target_rank*, or ``None`` if none does."""
    return next((row for row in standings if row.target_rank == target_rank), None)


def read_archon_xlsx(zip_bytes: bytes) -> bytes:
    """Return the raw bytes of ``archon.xlsx`` inside a TDA archive zip."""
    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        return zf.read("archon.xlsx")


def iter_tda_deck_texts(zip_bytes: bytes) -> Iterator[tuple[str, str]]:
    """Yield ``(filename, text)`` for every deck ``.txt`` entry in a TDA archive zip.

    Skips ``archon.xlsx`` (the only non-deck member observed in these archives).
    """
    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        for info in zf.infolist():
            if not info.filename.lower().endswith(".txt"):
                continue
            yield info.filename, zf.read(info).decode("utf-8")
